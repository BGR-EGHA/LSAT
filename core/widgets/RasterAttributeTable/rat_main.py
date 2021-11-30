import numpy as np
import numpy.lib.recfunctions as npf
import sys
import os
from osgeo import gdal
from osgeo import gdalconst
import logging
from openpyxl import Workbook
from core.libs.CustomFileDialog.CustomFileDialog import CustomFileDialog
from core.libs.GDAL_Libs.Layers import Raster
from core.widgets.RasterInfo.rasterInfo_main import RasterInfo
from core.uis.AttributeTable_ui.attributeTable_ui import Ui_AttributeTable
from core.uis.RasterAttributeTable_ui.AddField_ui import Ui_AddField
from PyQt5 import QtGui, QtCore, QtWidgets
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *


class AddField(QDialog):
    """
    This class controls the GUI for add field activity.
    """
    def __init__(self, parent=None):
        QDialog.__init__(self, parent)
        self.ui = Ui_AddField()
        self.ui.setupUi(self)
        self.setWindowIcon(QIcon(':/icons/Icons/AddField.png'))

    @pyqtSlot()
    def on_cancelPushButton_clicked(self):
        self.reject()

    @pyqtSlot()
    def on_applyPushButton_clicked(self):
        """
        This method controls the activities on apply button.
        :return: column name and column type specified in the GUI
        """
        global colName, colType
        if self.ui.fieldNameLineEdit.text() == "":
            QMessageBox.warning(
                self,
                self.tr("Warning"),
                self.tr("Field name required"),
                QMessageBox.Ok)
            return
        colName = self.ui.fieldNameLineEdit.text()
        colType = self.ui.fieldTypeComboBox.currentText()
        self.accept()
        return colName, colType


class RasterAttributeTable(QMainWindow):
    def __init__(self, raster_path, parent=None):
        QMainWindow.__init__(self, parent)
        # ui
        self.ui = Ui_AttributeTable()
        self.ui.setupUi(self)
        self.setWindowIcon(QIcon(':/icons/Icons/AttributeTable.png'))
        self.setWindowTitle(self.tr("Raster Attribute Table - {}").format(raster_path))
        self.RAT_status = True
        self.dialog = CustomFileDialog()
        self.raster_path = raster_path
        # Set up Toolbar and triggers
        self.actionSetEditable = QAction(
            QIcon(":/icons/Icons/File_Edit.png"),
            self.tr('Edit Attribute Table'),
            self)
        self.actionSetEditable.setCheckable(True)
        self.actionSetEditable.triggered.connect(self.setEditorMode)
        self.ui.toolBar.addAction(self.actionSetEditable)

        self.actionAddField = QAction(
            QIcon(":/icons/Icons/AddField.png"),
            self.tr('Add field'),
            self)
        self.actionAddField.triggered.connect(self.addColumn)
        self.actionAddField.setEnabled(False)
        self.ui.toolBar.addAction(self.actionAddField)

        self.actionDeleteField = QAction(
            QIcon(":/icons/Icons/DeleteField.png"),
            self.tr('Delete field'),
            self)
        self.actionDeleteField.triggered.connect(self.removeColumn)
        self.actionDeleteField.setEnabled(False)
        self.ui.toolBar.addAction(self.actionDeleteField)

        self.actionSaveEdits = QAction(
            QIcon(":/icons/Icons/SaveEdits.png"),
            self.tr('Save edits'),
            self)
        self.actionSaveEdits.triggered.connect(self.saveEdits)
        self.actionSaveEdits.setEnabled(False)
        self.ui.toolBar.addAction(self.actionSaveEdits)

        self.actionExport = QAction(
            QIcon(":/icons/Icons/OpenInExcel.png"),
            self.tr('Save RAT as xlsx-file'),
            self)
        self.actionExport.triggered.connect(self.exportTable)
        self.ui.toolBar.addAction(self.actionExport)

        self.ui.toolBar.addSeparator()
        actionGDALRasterInfo = QAction(
            QIcon(":/icons/Icons/raster_info.png"),
            self.tr('GDAL raster info'),
            self)
        actionGDALRasterInfo.triggered.connect(self.showRasterInfo)
        self.ui.toolBar.addAction(actionGDALRasterInfo)
        self.setEditorMode(False)  # initially disable editing

    def initRAT(self, state=False):
        """
        Gets called by setEditorMode and saveEdits.
        Reads the RAT and updates the the table.
        state defines if the table is editable after creating it.
        """
        self.raster = Raster(self.raster_path)
        self.layerName = os.path.basename(self.raster_path)
        self.rat = self.raster.band.GetDefaultRAT()
        if self.rat is None:
            bad_list = ["Float32", "CFloat32", "Float64", "CFloat64"]
            if self.raster.dataType in bad_list:
                logging.warning(
                    self.tr("{} has data type {}! A RAT cannot be generated for this data type!").format(
                        self.layerName, self.raster.dataType))
                self.RAT_status = False
                return
            else:
                reply = QMessageBox.question(
                    self,
                    self.tr("No RAT"),
                    self.tr("Selected dataset has no Attribute Table. Should a RAT be generated?"),
                    QMessageBox.Ok,
                    QMessageBox.Cancel)
                if reply == QMessageBox.Ok:
                    self.generateBasicRAT()
                    self.rat = self.raster.band.GetDefaultRAT()
                    self.names = self.getColNames()
                    self.table = self.rat2array()
                else:
                    self.RAT_status = False
                    self.close()
                    return
        self.table = self.rat2array()
        self.colNames = self.getColNames()
        self.model = TableModel(self.table.tolist(), self.colNames, editMode=state)
        self.ui.tableView.setModel(self.model)

    def generateBasicRAT(self):
        """
        create a basic RAT for the raster datset contaning value and count field
        :return: None
        """
        rat = gdal.RasterAttributeTable()
        values, counts = np.unique(self.raster.getArrayFromBand(), return_counts=True)
        rat.CreateColumn("OID", gdal.GFT_Integer, gdalconst.GFU_MinMax)
        rat.CreateColumn("VALUE", gdal.GFT_Real, gdalconst.GFU_MinMax)
        rat.CreateColumn("COUNT", gdal.GFT_Integer, gdalconst.GFU_MinMax)

        for i in range(len(values)):
            rat.SetValueAsInt(i, 0, int(i))
            rat.SetValueAsDouble(i, 1, float(values[i]))
            rat.SetValueAsInt(i, 2, int(counts[i]))
        self.raster.band.SetDefaultRAT(rat)
        self.raster.band.FlushCache()

    def setEditorMode(self, state: bool):
        """
        Controls the edit mode of the attribute table.
        :return: None
        """
        self.initRAT(state)
        self.actionAddField.setEnabled(state)
        self.actionDeleteField.setEnabled(state)
        self.actionSaveEdits.setEnabled(state)

    def closeEvent(self, event):
        """
        Closes the application and cleanup loaded data.
        :param event: Close Event
        :return: None
        """
        self.raster = None

    def exportTable(self):
        """
        Exports the table to a specified data format
        :return: None
        """
        self.dialog.saveExcelFile(os.path.dirname(self.raster_path))
        if self.dialog.exec_() == 1 and self.dialog.selectedFiles():
            if os.path.splitext(self.dialog.selectedFiles()[0])[1] != ".xlsx":
                fullFileName = self.dialog.selectedFiles()[0] + '.xlsx'
            else:
                fullFileName = self.dialog.selectedFiles()[0]
            headerNames = self.model.headerdata
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = self.tr("Table {}").format(os.path.basename(self.raster_path))

            for i, name in enumerate(headerNames, start=1):
                worksheet.cell(column=i, row=1, value=name)

            # values to table
            for i, row in enumerate(self.model.arraydata, start=1):
                for j, col in enumerate(row, start=1):
                    worksheet.cell(column=j, row=i + 1, value=col)
            # workbook.close()
            workbook.save(str(fullFileName))
            logging.info(
                self.tr("Table exported! Attribute table {} created.").format(fullFileName))

    def showRasterInfo(self):
        """
        Calls the raster info widget.
        :return: None
        """
        self.rasterInfo = RasterInfo(self.raster_path)
        self.rasterInfo.show()

    def saveEdits(self):
        """
        Writes the modifications into the GDAL raster attribute table.
        :return: None
        """
        headerNames = list(self.model.headerdata)
        table = list(self.model.arraydata)

        rat = gdal.RasterAttributeTable()

        for column in self.columns:
            if 'int' in str(column[1]):
                DTYPE = gdal.GFT_Integer
            elif 'float' in str(column[1]):
                DTYPE = gdal.GFT_Real
            elif 'U100' in str(column[1]):
                DTYPE = gdal.GFT_String
            else:
                pass
            rat.CreateColumn(str(column[0]), DTYPE, gdalconst.GFU_MinMax)

        for col in range(len(self.columns)):
            colType = rat.GetTypeOfCol(col)
            for row in range(len(table)):
                if colType == 0:
                    rat.SetValueAsInt(int(row), int(col), int(table[row][col]))
                elif colType == 1:
                    rat.SetValueAsDouble(row, col, table[row][col])
                elif colType == 2:
                    rat.SetValueAsString(row, col, str(table[row][col]))

        self.raster.band.SetDefaultRAT(rat)
        self.raster = None
        logging.info(self.tr("Edits saved."))
        self.initRAT(True)

    def removeColumn(self):
        """
        Removes current column from the table.
        :return: None
        """
        nonEditableFields = ["VALUE", "COUNT", "ID", "OID"]
        index = self.ui.tableView.currentIndex()
        if str(self.model.headerdata[index.column()]) in nonEditableFields:
            QMessageBox.warning(self, self.tr("Protected Field"), self.tr(
                "The field '{}' is a protected basic field and cannot be deleted!").format(str(self.model.headerdata[index.column()])))
            return
        else:
            msg = QMessageBox.question(self, self.tr("Confirm Delete Field"), self.tr(
                "Delete Field '{}'?").format(str(self.model.headerdata[index.column()])))
            if msg == QMessageBox.Yes:
                self.columns.pop(index.column())
                self.model.deleteColumn(index)
                self.ui.tableView.setCurrentIndex(QModelIndex())

    def addColumn(self):
        """
        Adds a column with specified type to the table.
        :return: None
        """
        dtype = {"Integer": 'int', "Double": 'float', "Text": 'U100'}

        dialog = AddField()
        if dialog.exec_() == 1:
            self.columns.append((str(colName), str(dtype[str(colType)])))
            if colName in self.model.headerdata:
                QMessageBox.warning(self,
                                    self.tr("Field already exists!"),
                                    self.tr("Please rename field!"))
                return
            # update qabstarcttable model
            self.model.insertColumn(str(colName), str(colType), QModelIndex())

    def rat2array(self):
        """
        Read a raster attribute table(rat) to numpy array
        :return: ndarray - structured numpy array
        """
        self.columns = []
        dtype = {0: 'int', 1: 'float', 2: 'U100'}
        for i in range(self.rat.GetColumnCount()):
            self.columns.append((str(self.rat.GetNameOfCol(i)).upper(),
                                str(dtype[self.rat.GetTypeOfCol(i)])))
        colNames = self.getColNames()
        table = np.zeros(shape=[self.rat.GetRowCount(), ], dtype=self.columns)
        for i in range(len(colNames)):
            array = self.rat.ReadAsArray(i).tolist()
            for row, element in enumerate(array):
                if self.columns[i][1] == "U100":
                    try:
                        table[row][i] = str(element, "utf8")
                    except BaseException:
                        table[row][i] = str(element, "cp1252")
                else:
                    table[row][i] = element
        return table

    def getColNames(self):
        """
        Return a list with column names of the raster attribute table
        :return: list with column names
        """
        names = []
        for i in range(self.rat.GetColumnCount()):
            names.append(str(self.rat.GetNameOfCol(i)))
        return names

class TableModel(QAbstractTableModel):
    def __init__(self, datain, headerdata, editMode=False, parent=None, *args):
        """
        datain: list of tupels or list of lists
        headerdata: a list of strings
        """
        QAbstractTableModel.__init__(self, parent, *args)
        # arraydata need to be a list of lists if setData should be applied (editable Table)
        # if the model should not be edited, one can pass list of tupels to prevent overriding.
        self.arraydata = [list(elem) for elem in datain]
        self.headerdata = headerdata
        self.editMode = editMode

    def rowCount(self, parent):
        return len(self.arraydata)

    def columnCount(self, parent):
        return len(self.headerdata)

    def deleteColumn(self, index):
        self.headerdata.remove(str(self.headerdata[index.column()]))
        for elem in self.arraydata:
            elem.pop(index.column())
        self.layoutChanged.emit()

    def insertColumn(self, name, type, index):
        self.headerdata.append(str(name))
        for elem in self.arraydata:
            if type == "Double":
                elem.append(0.0)
            elif type == "Integer":
                elem.append(0)
            else:
                elem.append("")
        self.layoutChanged.emit()

    def setData(self, index, value, role):
        self.arraydata[index.row()][index.column()] = value
        return True

    def data(self, index, role):
        if not index.isValid():
            return QVariant()
        elif role == Qt.TextAlignmentRole:
            return Qt.AlignCenter
        elif role == Qt.EditRole:
            return QVariant(self.arraydata[index.row()][index.column()])
        elif role != Qt.DisplayRole:
            return QVariant()
        return QVariant(self.arraydata[index.row()][index.column()])

    def headerData(self, col, orientation, role=Qt.DisplayRole):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return QVariant(self.headerdata[col])
        if role == Qt.TextAlignmentRole:
            return Qt.AlignCenter
        if role == Qt.EditRole:
            return Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.AlignCenter
        return QVariant()

    def flags(self, index):
        nonEditableFields = ["VALUE", "COUNT", "ID", "OID"]
        if self.editMode == True:
            if str(self.headerdata[index.column()]) in nonEditableFields:
                return Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.AlignCenter
            else:
                return Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.AlignCenter | Qt.ItemIsEditable
        else:
            return Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.AlignCenter
