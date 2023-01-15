# -*- coding: utf-8 -*-

import numpy as np
import sys
import os
import logging
from osgeo import ogr
from core.libs.CustomFileDialog.CustomFileDialog import CustomFileDialog
from core.uis.AttributeTable_ui.attributeTable_ui import Ui_AttributeTable
from core.widgets.FeatureInfo.featureInfo_main import FeatureInfo
from PyQt5 import QtCore
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from openpyxl import Workbook


class FeatureAttributeTable(QMainWindow):
    def __init__(self, feature_path, parent=None):
        QMainWindow.__init__(self, parent)
        self.ui = Ui_AttributeTable()
        self.ui.setupUi(self)
        self.setWindowIcon(QIcon(':/icons/Icons/File_Table.png'))
        self.featPath = feature_path
        self.setWindowTitle(self.tr("Attribute Table - {}").format(self.featPath))
        self.feature = ogr.Open(self.featPath)
        self.layer = self.feature.GetLayer()
        self.colNames = self.getColNames()
        self.colTypes = self.getColTypes()
        self.fieldCount = self.getFieldCount()
        self.table = self.fat2array()
        self.model = TableModel(self.table.tolist(), self.colNames)
        self.dialog = CustomFileDialog()
        self.ui.tableView.setModel(self.model)
        self.ui.tableView.setAlternatingRowColors(True)
        self.toolbar = self.ui.toolBar
        actionFeatureInfo = QAction(
            QIcon(":/icons/Icons/Properties_bw.png"),
            self.tr('Feature info'),
            self)
        actionFeatureInfo.triggered.connect(self.featureInfo)
        self.toolbar.addAction(actionFeatureInfo)

        actionExport = QAction(
            QIcon(":/icons/Icons/OpenInExcel.png"),
            self.tr('Save Attribute Table as xlsx-file'),
            self)
        actionExport.triggered.connect(self.exportTable)
        self.toolbar.addAction(actionExport)

    def closeEvent(self, event):
        if event:
            pass
        else:
            self.feature = None

    def featureInfo(self):
        self.featInfo = FeatureInfo(self.featPath)
        self.featInfo.show()

    def exportTable(self):
        """
        Exports the table to a specified data format
        :return: None
        """
        self.dialog.saveExcelFile(os.path.dirname(self.featPath))
        if self.dialog.exec_() == 1 and self.dialog.selectedFiles():
            if os.path.splitext(self.dialog.selectedFiles()[0])[1].lower() != ".xlsx":
                fullFileName = self.dialog.selectedFiles()[0] + '.xlsx'
            else:
                fullFileName = self.dialog.selectedFiles()[0]
            headerNames = self.table.dtype.names
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = self.tr("Table {}").format(os.path.basename(self.featPath))

            for i, name in enumerate(headerNames, start=1):
                worksheet.cell(column=i, row=1, value=name)

            # populate table
            for i, row in enumerate(self.table, start=1):
                for j, col in enumerate(row, start=1):
                    worksheet.cell(column=j, row=i + 1, value=col)
           # workbook.close()
            workbook.save(str(fullFileName))
            logging.info(
                self.tr("Table successfuly exported! Attribute table {} created.").format(fullFileName))

    def fat2array(self):
        columns = []
        dtype = {'Integer': 'U100',
                 'Integer64': 'U100',
                 'Double': 'float',
                 'Float': 'float',
                 'String': 'U100',
                 'StringList': 'U100',
                 'Text': 'U100',
                 'IntegerList': 'U100',
                 'Integer64List': 'U100',
                 'Real': "U100",
                 'RealList': "U100",
                 '(unknown)': 'U100',
                 'Binary': 'U100',
                 'Date': "U100",
                 'Time': "U100",
                 'DateTime': "U100"}

        for i, name in enumerate(self.colNames):
            try:
                DTYPE = dtype[str(self.colTypes[i])]
            except BaseException:
                DTYPE = "U100"

            columns.append((str(name), DTYPE))

        table = np.zeros(shape=(self.layer.GetFeatureCount(),), dtype=columns)
        for i, feat in enumerate(self.layer):
            for j in range(self.fieldCount):
                table[i][j] = feat.GetField(j)
        return table

    def getFieldCount(self):
        ldefn = self.layer.GetLayerDefn()
        count = ldefn.GetFieldCount()
        return count

    def getColNames(self):
        colNames = []
        ldefn = self.layer.GetLayerDefn()
        for i in range(ldefn.GetFieldCount()):
            fdefn = ldefn.GetFieldDefn(i)
            colNames.append(fdefn.name)
        return colNames

    def getColTypes(self):
        colTypes = []
        ldefn = self.layer.GetLayerDefn()
        for i in range(ldefn.GetFieldCount()):
            fdefn = ldefn.GetFieldDefn(i)
            colTypes.append(fdefn.GetFieldTypeName(i))
        return colTypes


class TableModel(QAbstractTableModel):
    def __init__(self, datain, headerdata, parent=None, *args):
        """ datain: a list of lists
            headerdata: a list of strings
        """
        QAbstractTableModel.__init__(self, parent, *args)
        self.arraydata = datain
        self.headerdata = headerdata

    def rowCount(self, parent):
        return len(self.arraydata)

    def columnCount(self, parent):
        return len(self.arraydata[0])

    def data(self, index, role):
        if not index.isValid():
            return QVariant()
        if role != Qt.DisplayRole:
            return QVariant()
        return QVariant(self.arraydata[index.row()][index.column()])

    def headerData(self, col, orientation, role):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return QVariant(self.headerdata[col])
        return QVariant()

    def flags(self, index):
        if index.column() is not None:
            return Qt.ItemIsEnabled | Qt.ItemIsSelectable
        return QtCore.QAbstractTableModel.flags(self, index)

    def sort(self, col, order):
        """
        Sort table by values in col. order defines if values are ascending or descending.
        Checks if all values are numbers before sorting, to avoid Errors like 9 > 111.
        """
        self.layoutAboutToBeChanged.emit()
        sortlist = [x[col].replace(".", "", 1) for x in self.arraydata]
        if all(list(map(str.isdigit, sortlist))):
            self.arraydata.sort(key=lambda x: float(x[col]), reverse=order)
        else:
            self.arraydata.sort(key=lambda x: x[col], reverse=order)
        self.layoutChanged.emit()
