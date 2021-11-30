# -*- coding: utf-8 -*-

import numpy as np
import itertools
import traceback
import os
import sys
import time
import logging
import math
from openpyxl import Workbook
from PyQt5 import QtGui, QtCore, QtWidgets
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from core.libs.CustomFileDialog.CustomFileDialog import CustomFileDialog
from core.libs.GDAL_Libs.Layers import Raster
from core.libs.LSAT_Messages.messages_main import Messenger
from core.widgets.RasterAttributeTable.rat_main import RasterAttributeTable
from core.uis.ResultViewer_ui.ResultsContingencyMatrix_ui import Ui_ContingencyMatrix
from core.uis.ResultViewer_ui.ResultsPairDetails_ui import Ui_ContingencyResults
from core.libs.LSAT_Messages.messages_main import Messenger


class ContingencyMatrix(QMainWindow):
    def __init__(self, projectLocation=None, fname=None, results=None, parent=None):
        QMainWindow.__init__(self, parent)
        self.ui = Ui_ContingencyMatrix()
        self.ui.setupUi(self)
        self.setWindowTitle(self.tr("Contingency - {}").format(str(fname)))
        self.setWindowIcon(QIcon(':/icons/Icons/Contingency_tab.png'))
        self.export2excelToolButton = QToolButton()
        self.export2excelToolButton.setIcon(QIcon(':/icons/Icons/OpenInExcel.png'))
        self.toolBar = QToolBar()
        self.addToolBar(self.toolBar)
        self.toolBar.addWidget(self.export2excelToolButton)
        self.export2excelToolButton.clicked.connect(self.exportTable)
        self.ui.tableView.doubleClicked.connect(self.showDetails)
        self.fileDialog = CustomFileDialog()
        self.projectLocation = projectLocation
        self.matrix_C = results["matrix_C"]
        self.matrix_CV = results["matrix_CV"]
        self.results = results["results"][0]

        self.ui.typeComboBox.activated.connect(self.updateMatrixView)
        self.updateMatrixView()

    def updateMatrixView(self):
        if self.ui.typeComboBox.currentText() == "Pearson's C":
            self.model = TableModel(self.matrix_C)
        else: # Cramer's V
            self.model = TableModel(self.matrix_CV)
        self.ui.tableView.setModel(self.model)
        self.ui.tableView.resizeColumnsToContents()

    def showDetails(self, index):
        """
        Gets called by double clicking a field in the table.
        Opens the Analysis Details window, if row != column (which would compare a raster against 
        itself).
        """
        if self.results is not None and index.row() != index.column():
            self.details = ContingencyResults(
                self.projectLocation, self.results[(index.row(), index.column())], parent=self)
            self.details.show()

    def exportTable(self):
        """
        Exports the table to a specified data format
        :return: None
        """
        self.fileDialog.saveExcelFile(self.projectLocation)
        if self.fileDialog.exec_() == 1 and self.fileDialog.selectedFiles():
            if os.path.splitext(self.fileDialog.selectedFiles()[0])[1] != ".xlsx":
                fullFileName = self.fileDialog.selectedFiles()[0] + '.xlsx'
            else:
                fullFileName = self.fileDialog.selectedFiles()[0]
            if os.path.exists(fullFileName):
                ret = self.msg.WarningOverwriteFile()
                if ret == QMessageBox.Cancel:
                    return
            header_h = self.model.header
            header_v = self.model.header

            workbook = Workbook()

            tableList = [("Pearson's C", self.matrix_C), (" Cramer's V", self.matrix_CV)]
            for i, elem in enumerate(tableList):
                if i == 0:
                    worksheet = workbook.active
                    worksheet.title = elem[0]
                else:
                    worksheet = workbook.create_sheet(elem[0])

                for i, name in enumerate(header_v, start=1):
                    worksheet.cell(column=i + 1, row=1, value=name)
                for i, name in enumerate(header_h, start=1):
                    worksheet.cell(column=1, row=i + 1, value=name)
                # populate table
                for i, row in enumerate(elem[1], start=1):
                    for j, col in enumerate(row, start=1):
                        worksheet.cell(column=j + 1, row=i + 1, value=col)
            # workbook.close()
            workbook.save(str(fullFileName))
            logging.info(
                self.tr("Table exported! Attribute table {} created.").format(fullFileName))


class ContingencyResults(QMainWindow):
    def __init__(self, projectLocation=None, results=None, parent=None):
        QMainWindow.__init__(self, parent)
        self.ui = Ui_ContingencyResults()
        self.ui.setupUi(self)
        self.setWindowTitle(self.tr("Analysis Details"))
        self.fileDialog = CustomFileDialog()
        self.toolBar = QToolBar()
        self.addToolBar(self.toolBar)
        self.msg = Messenger()

        self.export2excelToolButton = QToolButton()
        self.export2excelToolButton.setIcon(QIcon(':/icons/Icons/OpenInExcel.png'))
        self.toolBar.addWidget(self.export2excelToolButton)

        self.export2excelToolButton.clicked.connect(self.exportTable)

        self.projectLocation = projectLocation
        self.C = results[0]
        self.CV = results[1]
        self.Phi = results[2]
        self.X2 = results[3]
        self.cross_table = results[4]
        self.header_v = results[5]
        self.header_h = results[6]

        model_Phi = TableModel(self.Phi, (self.header_v, self.header_h), phi=True)
        self.ui.phiTableView.setModel(model_Phi)
        model_X2 = TableModel(self.X2, (self.header_v, self.header_h))
        self.ui.chi2TableView.setModel(model_X2)
        model_crossTab = TableModel(self.cross_table, (self.header_v, self.header_h))
        self.ui.frequencyTableView.setModel(model_crossTab)

    def exportTable(self):
        """
        Exports the table to a specified data format
        :return: None
        """
        self.fileDialog.saveExcelFile(self.projectLocation)
        if self.fileDialog.exec_() == 1 and self.fileDialog.selectedFiles():
            if os.path.splitext(self.fileDialog.selectedFiles()[0])[1] != ".xlsx":
                fullFileName = self.fileDialog.selectedFiles()[0] + '.xlsx'
            else:
                fullFileName = self.fileDialog.selectedFiles()[0]
            if os.path.exists(fullFileName):
                ret = self.msg.WarningOverwriteFile()
                if ret == QMessageBox.Cancel:
                    return
            header_h = self.header_h
            header_v = self.header_v
            workbook = Workbook()

            tableList = [("Phi", self.Phi), ("Chi2", self.X2),
                         ("Frequency table", self.cross_table)]
            for i, elem in enumerate(tableList):
                if i == 0:
                    worksheet = workbook.active
                    worksheet.title = elem[0]
                else:
                    worksheet = workbook.create_sheet(elem[0])

                for i, name in enumerate(header_v, start=1):
                    worksheet.cell(column=i + 1, row=1, value=name)
                for i, name in enumerate(header_h, start=1):
                    worksheet.cell(column=1, row=i + 1, value=name)
                # populate table
                for i, row in enumerate(elem[1], start=1):
                    for j, col in enumerate(row, start=1):
                        worksheet.cell(column=j + 1, row=i + 1, value=col)
            # workbook.close()
            workbook.save(str(fullFileName))
            logging.info(
                self.tr("Table exported! Attribute table {} created.").format(fullFileName))


class TableModel(QAbstractTableModel):
    def __init__(self, indata, header=None, phi=False, parent=None, *args):
        QAbstractTableModel.__init__(self, parent)
        self.indata = indata
        self.phi = phi
        if header is None:
            self.header = indata.dtype.names
            self.tabMode = True
        else:
            self.header_v = header[0]
            self.header_h = header[1]
            self.tabMode = False

    def rowCount(self, QModelIndex_parent=None, *args, **kwargs):
        return self.indata.shape[0]

    def columnCount(self, QModelIndex_parent=None, *args, **kwargs):
        if self.tabMode:
            return len(self.header)
        else:
            return len(self.header_h)

    def data(self, index, int_role=QtCore.Qt.DisplayRole):
        if int_role == QtCore.Qt.DisplayRole:
            i = index.row()
            j = index.column()
            return '{:10.2f}'.format(self.indata[i][j])
        if int_role == Qt.TextAlignmentRole:
            return Qt.AlignCenter
        if self.tabMode:
            if int_role == Qt.BackgroundRole and index.column() == index.row():
                return QBrush(Qt.black)
            if int_role == Qt.BackgroundRole and index.column() < index.row():
                return QBrush(Qt.lightGray)
        if self.tabMode == False and self.phi and int_role == Qt.BackgroundRole:
            i = index.row()
            j = index.column()
            if self.indata[i][j] >= 0.5:
                return QBrush(Qt.red)
            elif self.indata[i][j] > 0.3 and self.indata[i][j] < 0.5:
                return QBrush(Qt.yellow)
            else:
                return QBrush(Qt.green)

        else:
            return QtCore.QVariant()

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if self.tabMode:
            if role == Qt.DisplayRole and orientation == Qt.Horizontal:
                return self.header[section]
            if role == Qt.DisplayRole and orientation == Qt.Vertical:
                return self.header[section]
        else:
            if role == Qt.DisplayRole and orientation == Qt.Horizontal:
                return self.header_h[section]
            if role == Qt.DisplayRole and orientation == Qt.Vertical:
                return self.header_v[section]
        return QAbstractTableModel.headerData(self, section, orientation, role)
