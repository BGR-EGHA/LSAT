# -*- coding: utf-8 -*-

from osgeo import gdal, ogr, osr, gdalconst
import sys
import os
import math
import time
import numpy as np
from openpyxl import Workbook
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
import logging
import traceback
from shutil import copy2

from core.libs.GDAL_Libs.Layers import Raster, Feature, RasterLayer
from core.uis.ResultViewer_ui.tableViewer_ui import Ui_TableViewWidget
from core.libs.CustomFileDialog.CustomFileDialog import CustomFileDialog
from core.libs.LSAT_Messages.messages_main import Messenger


class TableModel(QAbstractTableModel):
    def __init__(self, indata, v_header, parent=None, *args):
        QAbstractTableModel.__init__(self, parent)
        self.indata = indata
        self.header = indata.dtype.names
        self.dtypes = [self.indata.dtype[i] for i in range(len(self.header))]
        self.verticalHeader = v_header

    def rowCount(self, QModelIndex_parent=None, *args, **kwargs):
        return self.indata.shape[0]

    def columnCount(self, QModelIndex_parent=None, *args, **kwargs):
        return len(self.header)

    def data(self, index, role=QtCore.Qt.DisplayRole):
        if role == QtCore.Qt.DisplayRole:
            i = index.row()
            j = index.column()
            if "int" in str(self.dtypes[j]):
                return '{0}'.format(self.indata[i][j])
            else:
                return '{:.5f}'.format(self.indata[i][j])

        if not index.isValid():
            return QVariant()

        elif role == Qt.TextAlignmentRole:
            return Qt.AlignCenter

        elif role != Qt.DisplayRole:
            return QVariant()
        else:
            return QVariant()

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role == Qt.DisplayRole and orientation == Qt.Horizontal:
            return self.header[section]
        if role == Qt.DisplayRole and orientation == Qt.Vertical:
            return QVariant(self.verticalHeader[section])
        return QAbstractTableModel.headerData(self, section, orientation, role)


class TableViewer(QMainWindow):
    def __init__(self, projectLocation, layerName, result, analysis, parent=None):
        QWidget.__init__(self, parent)
        self.ui = Ui_TableViewWidget()
        self.ui.setupUi(self)
        self.setWindowIcon(QIcon(':/icons/Icons/Chart_Bar_Big.png'))
        self.fileDialog = CustomFileDialog()
        self.msg = Messenger()

        self.tableView = self.ui.tableView
        self.progressBar = QProgressBar()
        self.ui.statusbar.addPermanentWidget(self.progressBar, 0)

        self.projectLocation = projectLocation
        self.resultsLocation = os.path.join(projectLocation, "results", analysis)
        self.weightedLayersLocation = os.path.join(self.resultsLocation, "rasters")
        self.tableLocation = os.path.join(self.resultsLocation, "tables")

        self.result = result
        self.layerName = os.path.basename(self.result["source"][0]).split(".")[0]

        self.table = self.result["tab"]
        self.tableView.setStyleSheet("QHeaderView::section { background-color:#b7cbeb }")

        self.actionExport = QAction(
            QIcon(":/icons/Icons/OpenInExcel.png"),
            self.tr('Export to Excel'),
            self)
        self.actionExport.triggered.connect(self.exportTable)
        self.ui.toolBar.addAction(self.actionExport)

        self.actionWriteRaster = QAction(
            QtGui.QIcon(':/icons/Icons/WriteRaster.png'),
            self.tr('Write Raster from Column'),
            self)
        self.actionWriteRaster.triggered.connect(self.writeRasterFromColumn)
        self.ui.toolBar.addAction(self.actionWriteRaster)

        # combo label field takes columns of the raster attribute table and
        # updates the table vertical header anytime the value is changed
        self.comboLabelField = QComboBox()
        self.comboLabelField.activated.connect(self.updateView)
        self.comboLabelField.setEnabled(False)
        self.ui.toolBar.addWidget(self.comboLabelField)
        self.getLabelNames()  # Gets labels based on the rasters rat table
        if self.labels != ["Unique Raster Values"]:
            headerValues = self.ratTable[str(self.comboLabelField.currentText())]
            headerList = headerValues.tolist()
        else:  # We could not read the RAT and use unique Raster values
            headerList = self.uniquevalues
        if -9999 in headerList:
            headerList.remove(-9999)
        self.model = TableModel(self.table, headerList)
        self.tableView.setModel(self.model)

    def updateView(self):
        """
        Update the table any time when the value in label combo box
        has changed.
        :return: None
        """
        if self.labels is not None:
            headerValues = self.ratTable[str(self.comboLabelField.currentText())]
            headerList = headerValues.tolist()
            if -9999 in headerList:
                headerList.remove(-9999)
            self.model = TableModel(self.table, headerList)
            self.tableView.setModel(self.model)
            self.tableView.updateGeometries()

    def getLabelNames(self):
        """
        This method populates the label combo box with
        column names contained in the raster attribute table.
        :return: list of column names in the rat
        """
        raster = Raster(str(self.result["source"][0]))
        if raster.rat is not None:
            self.ratTable = raster.rat2array()
            self.comboLabelField.setEnabled(True)
            self.labels = self.ratTable.dtype.names
            for label in self.labels:
                self.comboLabelField.insertItem(0, str(label))
        else:
            logging.warning(
                self.tr("Could not read raster attribute table of {}").format(
                    self.result['source'][0]))
            self.labels = ["Unique Raster Values"]
            self.uniquevalues = raster.uniqueValues.tolist()
            self.comboLabelField.setEnabled(False)
        for label in self.labels:
            self.comboLabelField.insertItem(0, str(label))

    def writeRasterFromColumn(self):
        """
        This method writes a Tiff raster from a selected table column.
        :return: None
        """
        try:
            selection = self.tableView.selectionModel().selectedColumns()
            if len(selection) == 0:
                QMessageBox.warning(self, self.tr("No column selected"), self.tr(
                    "Select a column to to write values to raster"))
                return
            if len(selection) > 1:
                QMessageBox.warning(self, self.tr("Multiple Selection"),
                                    self.tr("Select only one column!"))
                return
            sourceResultPath = os.path.join(str(self.tableLocation),
                                            '{}_tab.npz'.format(str(self.layerName)))
            idx = selection[0].column()
            type = self.model.dtypes[idx]
            if "int" in str(type):
                dataType = gdal.GDT_Int16
            elif "float" in str(type):
                dataType = gdal.GDT_Float32

            outRasterList = {
                u"Class": '_cl',
                u"Landslides": '_land',
                u"W_POS": '_pos',
                u"VAR_POS": '_varp',
                u"W_NEG": '_neg',
                u"VAR_NEG": '_varn',
                u"Variance": '_var',
                u"Contrast": '_C',
                u"Weight": '_woe',
                u"Posterior": '_pp',
                u"sPost": '_spp',
                u"Expected": '_xpc',
                u"sExpec": '_sxpc'}

            # Get data from table column
            value_list = [self.model.indata[i][idx] for i in range(len(self.model.indata))]

            self.fileDialog.saveRasterFile(str(self.weightedLayersLocation), str(
                self.layerName) + str(outRasterList[str(self.model.header[idx])]))
            if self.fileDialog.exec_() == 1:
                for filename in self.fileDialog.selectedFiles():
                    if filename is None or str(filename) == "":
                        return
                    else:
                        if str(filename).endswith(".tif"):
                            basename = os.path.basename(str(filename)).split(".")[0]
                        else:
                            basename = os.path.basename(str(filename))

                        # self.outTablePath = os.path.join(str(os.path.dirname(), basename +
                        # '_tab.npz')
                        self.outRasterPath = str(filename)
                        # Get reference raster (source)
                        raster = Raster(self.result["source"][0])
                        array = raster.getArrayFromBand()
                        uniques = np.unique(array)
                        raster_values = uniques[np.where(uniques != raster.nodata)]

                        # Create out raster
                        NoData_value = -9999.0
                        driver = gdal.GetDriverByName('GTiff')
                        if not self.outRasterPath.lower().endswith(".tif"):
                            self.outRasterPath = self.outRasterPath + ".tif"
                        outRaster = driver.Create(
                            self.outRasterPath, raster.cols, raster.rows, 1, gdal.GDT_Float32)
                        outRaster.SetProjection(raster.proj)
                        outRaster.SetGeoTransform(raster.geoTrans)
                        band = outRaster.GetRasterBand(1).SetNoDataValue(NoData_value)

                        outArray = np.ones_like(array) * -9999.0
                        for i, elem in enumerate(raster_values):
                            outArray[np.where(np.equal(array, elem))] = value_list[i]
                        outRaster.GetRasterBand(1).WriteArray(outArray)
                        outRaster.GetRasterBand(1).ComputeStatistics(False)
                        outRaster = None
                        logging.info(
                            self.tr("Output raster {} created.").format(
                                self.outRasterPath))
            outRaster = None
        except BaseException:
            tb = traceback.format_exc()
            logging.error(tb)

    def exportTable(self):
        """
        Exports the table to a specified data format
        :return: None
        """
        self.fileDialog.saveExcelFile(self.projectLocation)
        if self.fileDialog.exec_() == 1 and self.fileDialog.selectedFiles():
            if os.path.splitext(self.fileDialog.selectedFiles()[0])[1].lower() != ".xlsx":
                fullFileName = self.fileDialog.selectedFiles()[0] + '.xlsx'
            else:
                fullFileName = self.fileDialog.selectedFiles()[0]
            if os.path.exists(fullFileName):
                ret = self.msg.WarningOverwriteFile()
                if ret == QMessageBox.Cancel:
                    return
            headerNames = self.model.header
            classNames = self.model.verticalHeader
            workbook = Workbook()
            worksheet = workbook.active

            worksheet.title = self.tr("Table").format(self.layerName)

            for i, name in enumerate(headerNames, start=1):
                worksheet.cell(column=i + 1, row=1, value=name)
            for i, name in enumerate(classNames, start=1):
                worksheet.cell(column=1, row=i + 1, value=name)
            # populate table
            for i, row in enumerate(self.model.indata, start=1):
                for j, col in enumerate(row, start=1):
                    worksheet.cell(column=j + 1, row=i + 1, value=col)
            # workbook.close()
            workbook.save(str(fullFileName))
            logging.info(
                self.tr("Table exported! Attribute table {} created.").format(fullFileName))
